#!/usr/bin/env python3
"""
Build audit cash book with personal mobile loans stripped.

Keeps as business LOAN (per user):
  TCL CREDIT, JACKFRUIT, PREMIER CREDIT, MYCREDIT / MY CREDIT, UNI LIMITED

Also keeps bank-style loan repayments and non-LOAN expense rows.
Removes personal/digital mobile loan lines from the LOAN column.

Reuses salary insert + live TOTAL / Summary link formulas.

Output:
  C:\\Users\\brian\\Downloads\\down\\cash-book-2025-updated-audit-no-personal-mobile.xlsx
"""
from __future__ import annotations

import re
import sys
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

YEAR = 2025
SRC = Path(r"C:\Users\brian\Downloads\down\cash-book-2025-updated.xlsx")
OUT = Path(
    r"C:\Users\brian\Downloads\down\cash-book-2025-updated-audit-no-personal-mobile.xlsx"
)

DETAIL_SHEET = "Cash Book 2025"
TURNOVER = 24_767_320
SALARY_TOTAL = 5_769_461

MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
COLUMNS = [
    "FOOD", "Stationary", "Textbooks", "EXAMS", "Uniform", "Communication",
    "Office Exp", "Water", "General repair", "Furniture", "Medical",
    "Service provider", "Fuel", "NTSA", "Trash", "Colnet", "Construction",
    "Vehicle Repairs", "Transport", "Electricity", "Labour", "Donation",
    "Advance tax", "Insurance", "Lisence", "Assets", "LOAN", "Salary",
    "Valuation & Inspection", "ACTIVITIES", "NSSF", "SHA", "PAYE", "NITA",
    "Housing", "Rent",
]
HEADERS = ["DATE", "Voucher No.", "ITEM", "AMOUNT"] + COLUMNS
LOAN_COL = HEADERS.index("LOAN")  # 0-based
SALARY_COL = HEADERS.index("Salary") + 1
NCOLS = len(HEADERS)

# Business lenders to KEEP under LOAN
KEEP_BUSINESS_RE = re.compile(
    r"tcl\s*credit|jackfruit|premier\s*credit|my\s*credit|mycredit|uni\s*limited",
    re.I,
)

# Personal / digital mobile loans (and misc personal lines parked in LOAN) to REMOVE
STRIP_PERSONAL_RE = re.compile(
    r"|".join(
        [
            r"\bzenka\b",
            r"kcb\s*mpesa",
            r"aventus",
            r"azura",
            r"\btimiza\b",
            r"branch\s*micro",
            r"\btala\b",
            r"m-?shwari",
            r"fuliza",
            r"tingg",
            r"cellulant",
            r"signalwave",
            r"\bokoa\b",
            r"hfm\s*invest",
            r"term\s*loan\s*bridge",
            r"loan\s*account\s*payments",
            r"loan\s*recovery\s*for",
            r"education\s*expenses",
            r"^utilities$",
            r"goods\s*and\s*services",
            r"^fuel\s*expenses$",
        ]
    ),
    re.I,
)


def parse_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def month_of_row(date_v, vch) -> int | None:
    d = parse_date(date_v)
    if d:
        return d.month
    v = str(vch or "").strip().upper()
    for prefix in ("E-", "SAL-", "A-", "F-"):
        if v.startswith(prefix):
            v = v[len(prefix):]
            break
    if len(v) >= 3 and v[:3] in MONTHS:
        return MONTHS.index(v[:3]) + 1
    return None


def salary_splits(total: int = SALARY_TOTAL) -> list[int]:
    base = total // 12
    rem = total - base * 12
    amounts = [base] * 12
    amounts[11] += rem
    return amounts


def loan_amount(cells: list) -> float:
    try:
        return float(cells[LOAN_COL] or 0)
    except (TypeError, ValueError):
        return 0.0


def is_personal_mobile_loan(cells: list) -> bool:
    """True if this detail row should be stripped as personal mobile loan."""
    if loan_amount(cells) <= 0:
        return False
    item = str(cells[2] or "").strip()
    if KEEP_BUSINESS_RE.search(item):
        return False
    if STRIP_PERSONAL_RE.search(item):
        return True
    return False


def extract_detail_rows_with_strip(path: Path) -> tuple[dict[int, list[list]], dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[DETAIL_SHEET]
    by_month: dict[int, list[list]] = defaultdict(list)
    stripped = Counter()
    stripped_total = 0.0
    stripped_rows = 0
    kept = 0

    for row in ws.iter_rows(min_row=1, max_col=NCOLS, values_only=True):
        cells = list(row[:NCOLS])
        label = str(cells[0] or "").strip().upper()
        if label in ("DATE", "TOTAL"):
            continue
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        try:
            amt = float(cells[3] or 0)
        except (TypeError, ValueError):
            continue
        if not amt:
            continue
        m = month_of_row(cells[0], cells[1])
        if not m:
            continue
        while len(cells) < NCOLS:
            cells.append(None)
        cells = cells[:NCOLS]

        if is_personal_mobile_loan(cells):
            item = str(cells[2] or "").strip() or "(blank)"
            la = loan_amount(cells)
            stripped[item] += la
            stripped_total += la
            stripped_rows += 1
            continue

        by_month[m].append(cells)
        kept += 1

    wb.close()
    return by_month, {
        "kept": kept,
        "stripped_rows": stripped_rows,
        "stripped_total": stripped_total,
        "stripped_by_item": stripped,
    }


def write_workbook(by_month: dict[int, list[list]], path: Path) -> dict:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DETAIL_SHEET

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E1F2")
    int_fmt = "#,##0"
    profit_fill = PatternFill("solid", fgColor="FFF2CC")

    splits = salary_splits()
    month_total_rows: dict[int, int] = {}

    row_num = 1
    for m in range(1, 13):
        entries = list(by_month.get(m, []))

        last_day = monthrange(YEAR, m)[1]
        sal_amt = splits[m - 1]
        sal_row = [None] * NCOLS
        sal_row[0] = date(YEAR, m, last_day)
        sal_row[1] = f"SAL-{MONTHS[m - 1]}"
        sal_row[2] = "Salaries & wages"
        sal_row[3] = sal_amt
        sal_row[SALARY_COL - 1] = sal_amt
        entries.append(sal_row)

        for i, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_num, column=i, value=h)
            cell.font = bold
            cell.fill = fill
        row_num += 1
        first = row_num

        for line in entries:
            for i, val in enumerate(line, 1):
                cell = ws.cell(row=row_num, column=i, value=val)
                if i >= 4 and isinstance(val, (int, float)) and val:
                    cell.number_format = int_fmt
            row_num += 1

        last = row_num - 1
        ws.cell(row=row_num, column=1, value="TOTAL").font = bold
        for col_i in range(4, NCOLS + 1):
            letter = get_column_letter(col_i)
            cell = ws.cell(
                row=row_num,
                column=col_i,
                value=f"=SUM({letter}{first}:{letter}{last})",
            )
            cell.font = bold
            cell.number_format = int_fmt
        month_total_rows[m] = row_num
        row_num += 2

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"EXPENSES SUMMARY YEAR {YEAR}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"] = "Category"
    ws2["A2"].font = bold
    ws2["A2"].fill = fill
    for i, mon in enumerate(MONTHS, 2):
        cell = ws2.cell(row=2, column=i, value=mon)
        cell.font = bold
        cell.fill = fill
    cell = ws2.cell(row=2, column=14, value="TOTAL")
    cell.font = bold
    cell.fill = fill

    detail_ref = f"'{DETAIL_SHEET}'"
    cat_start = 3
    for ci, cat in enumerate(COLUMNS):
        r = cat_start + ci
        ws2.cell(row=r, column=1, value=cat)
        cat_letter = get_column_letter(5 + ci)
        for m in range(1, 13):
            tot_row = month_total_rows[m]
            cell = ws2.cell(
                row=r,
                column=m + 1,
                value=f"={detail_ref}!{cat_letter}{tot_row}",
            )
            cell.number_format = int_fmt
        cell = ws2.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
        cell.number_format = int_fmt

    cat_end = cat_start + len(COLUMNS) - 1
    gt_row = cat_end + 1
    ws2.cell(row=gt_row, column=1, value="GRAND TOTAL").font = bold
    for c in range(2, 15):
        letter = get_column_letter(c)
        cell = ws2.cell(
            row=gt_row,
            column=c,
            value=f"=SUM({letter}{cat_start}:{letter}{cat_end})",
        )
        cell.font = bold
        cell.number_format = int_fmt

    ws2.cell(row=cat_end, column=17, value="25k per month anything above")

    p0 = gt_row + 2
    ws2.cell(row=p0, column=1, value="P&L (linked)").font = Font(bold=True, size=12)
    ws2.cell(row=p0 + 1, column=1, value="Turnover (TB)")
    cell = ws2.cell(row=p0 + 1, column=2, value=TURNOVER)
    cell.number_format = int_fmt
    cell.fill = profit_fill
    ws2.cell(row=p0 + 2, column=1, value="Total expenses (cash book)")
    cell = ws2.cell(row=p0 + 2, column=2, value=f"=N{gt_row}")
    cell.number_format = int_fmt
    cell.fill = profit_fill
    ws2.cell(row=p0 + 3, column=1, value="Profit / (Loss)")
    cell = ws2.cell(row=p0 + 3, column=2, value=f"=B{p0 + 1}-B{p0 + 2}")
    cell.font = bold
    cell.number_format = int_fmt
    cell.fill = profit_fill
    ws2.cell(
        row=p0 + 5,
        column=1,
        value=(
            "Personal mobile loans removed from LOAN. Kept business: "
            "TCL CREDIT, JACKFRUIT, PREMIER CREDIT, MYCREDIT, UNI LIMITED "
            "(plus bank loan repayments). Salary includes TB Salaries & wages-NITA."
        ),
    )
    ws2.merge_cells(start_row=p0 + 5, start_column=1, end_row=p0 + 5, end_column=8)
    ws2.cell(row=p0 + 5, column=1).alignment = Alignment(wrap_text=True)

    ws2.column_dimensions["A"].width = 28
    for col in range(2, 15):
        ws2.column_dimensions[get_column_letter(col)].width = 11

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return {"month_total_rows": month_total_rows, "grand_total_row": gt_row}


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source: {SRC}")

    by_month, stats = extract_detail_rows_with_strip(SRC)
    if stats["kept"] == 0:
        raise SystemExit("No detail rows left after filter")

    write_workbook(by_month, OUT)

    base = sum(float(line[3] or 0) for rows in by_month.values() for line in rows)
    expected_expenses = int(base) + SALARY_TOTAL
    profit = TURNOVER - expected_expenses

    print("Cash book written (personal mobile loans stripped)")
    print(f"  {OUT}")
    print(f"  Rows kept: {stats['kept']}")
    print(f"  Rows stripped: {stats['stripped_rows']}  amount={stats['stripped_total']:,.0f}")
    print("  Stripped by item:")
    for item, amt in stats["stripped_by_item"].most_common():
        print(f"    {amt:>12,.0f}  {item}")
    print(f"  Expected expenses (incl salary): {expected_expenses:,}")
    print(f"  Turnover: {TURNOVER:,}")
    print(f"  Profit/(Loss): {profit:,}")


if __name__ == "__main__":
    main()
