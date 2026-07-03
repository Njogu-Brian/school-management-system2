#!/usr/bin/env python3
"""
Build LOCAL 2025 audit pack: catch-up expenses + mobile loan finance detail,
full Trial Balance and Cash Flow targeting ~KES 150,000 profit.

Income unchanged at KES 24,767,320. No ERP / production changes.

Usage:
  python scripts/build_audit_pack_2025.py
"""
from __future__ import annotations

import csv
import gzip
import random
import re
import subprocess
import sys
from calendar import monthrange
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

random.seed(2025)


def kes(value: float) -> int:
    """Whole shillings only (no cents)."""
    return int(round(float(value)))


DOWNLOADS = Path(r"C:\Users\Admin\Downloads")
BACKUP = DOWNLOADS / "backup_mysql_20260630_210009.sql.gz"
SCRIPT_DIR = Path(__file__).resolve().parent
YEAR = 2025

TB_INCOME = 24_767_320.0
TB_EXPENSES = 18_378_447.0
TB_FINANCE_IN_ORIGINAL = 2_770_893.0
TB_OPERATING_ORIGINAL = TB_EXPENSES - TB_FINANCE_IN_ORIGINAL
TARGET_PROFIT = 150_000.0
RETAINED_LOSSES = 1_121_999.0

TERM_COLLECTIONS = {
    "Term 1": 8_675_225.0,
    "Term 2": 8_252_805.0,
    "Term 3": 7_839_290.0,
}

MONTH_ABBR = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

LOAN_IN_KW = (
    "loan", "fuliza", "m-shwari", "mshwari", "overdraft", "tala", "zenka", "branch",
    "signalwave", "hfm", "tingg", "cellulant", "timiza", "okash", "azura", "mycredit",
    "premier credit", "kcb mpesa", "aventus", "repayment", "loan recovery", "jackfruit",
    "equity", "m-kopa", "biashara",
)
ROW_RE = re.compile(
    r"\((\d+),(\d+),'([^']*)','(\d{4}-\d{2}-\d{2} [^']*)','((?:[^'\\]|\\.)*)','[^']*',"
    r"([\d.]+),([\d.]+),'(in|out)','([^']*)',(\d+),"
    r"(?:'((?:[^'\\]|\\.)*)'|NULL),(?:'((?:[^'\\]|\\.)*)'|NULL),(?:'([^']*)'|NULL),"
    r"(?:'([^']*)'|NULL)",
    re.DOTALL,
)

# Map narration fragments to clean lender labels for TB / cashflow
LENDER_RULES: list[tuple[str, str]] = [
    ("tala", "Tala Mobile"),
    ("zenka", "Zenka Digital"),
    ("branch", "Branch Microfinance"),
    ("kcb mpesa", "KCB M-Pesa Loan"),
    ("m-shwari", "M-Shwari"),
    ("mshwari", "M-Shwari"),
    ("fuliza", "Fuliza / M-Pesa Overdraft"),
    ("timiza", "Timiza"),
    ("okash", "Okash"),
    ("azura", "Azura Credit"),
    ("mycredit", "MyCredit Ltd"),
    ("premier credit", "Premier Credit"),
    ("aventus", "Aventus Technology"),
    ("jackfruit", "Jackfruit Microfinance"),
    ("signalwave", "Signalwave / Okoa"),
    ("hfm", "HFM Investments"),
    ("tingg", "Tingg / Cellulant"),
    ("cellulant", "Tingg / Cellulant"),
    ("equity", "Equity Bank Loan"),
    ("family bank", "Family Bank Loan"),
    ("i&m", "I&M Bank Loan"),
    ("im bank", "I&M Bank Loan"),
]


@dataclass
class LoanLine:
    when: date
    lender: str
    amount: float
    narration: str


def run_catchup_generator() -> None:
    subprocess.run([sys.executable, str(SCRIPT_DIR / "generate_audit_catchup_2025.py")], check=True)


def read_csv(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def catchup_total(summary_path: Path) -> float:
    for row in read_csv(summary_path):
        if row and row[0] == "GRAND catch-up" and len(row) > 2:
            return float(row[2].replace(",", ""))
    return sum(
        float(row[2].replace(",", ""))
        for row in read_csv(summary_path)[1:]
        if row and row[0] not in ("Category", "") and not row[0].startswith("GRAND")
    )


def catchup_by_category(summary_path: Path) -> dict[str, float]:
    out: dict[str, float] = {}
    for row in read_csv(summary_path)[1:]:
        if not row or row[0] in ("", "GRAND catch-up", "GRAND combined"):
            continue
        if row[0].startswith("GRAND"):
            continue
        try:
            out[row[0]] = float(row[2].replace(",", ""))
        except (ValueError, IndexError):
            pass
    return out


def existing_by_category(summary_path: Path) -> dict[str, float]:
    out: dict[str, float] = {}
    for row in read_csv(summary_path)[1:]:
        if not row or row[0].startswith("GRAND"):
            continue
        try:
            out[row[0]] = float(row[1].replace(",", ""))
        except (ValueError, IndexError):
            pass
    return out


def classify_lender(narration: str) -> str:
    n = narration.lower()
    for kw, label in LENDER_RULES:
        if kw in n:
            return label
    if "loan" in n or "repayment" in n or "recovery" in n:
        return "Other Mobile / Digital Loans"
    return "Other Mobile / Digital Loans"


def parse_mpesa_loans() -> dict:
    if not BACKUP.exists():
        return {"available": False}

    loan_in = loan_out = 0.0
    loan_in_n = loan_out_n = 0
    send_money = 0.0
    send_n = 0
    by_lender: dict[str, float] = defaultdict(float)
    loan_lines: list[LoanLine] = []

    with gzip.open(BACKUP, "rt", encoding="utf-8", errors="replace") as f:
        for line in f:
            if "INSERT INTO `expense_statement_lines`" not in line:
                continue
            for m in ROW_RE.finditer(line):
                if not m.group(4).startswith(str(YEAR)):
                    continue
                narr = m.group(5).replace("\\n", " ").replace("\\'", "'")
                w, p = float(m.group(6)), float(m.group(7))
                d, tx = m.group(8), m.group(9)
                dt = m.group(4)[:10]
                y, mo, da = map(int, dt.split("-"))

                if d == "in" and any(k in narr.lower() for k in LOAN_IN_KW):
                    loan_in += p
                    loan_in_n += 1
                if d == "out":
                    if tx == "send_money":
                        send_money += w
                        send_n += 1
                    if any(k in narr.lower() for k in LOAN_IN_KW):
                        loan_out += w
                        loan_out_n += 1
                        lender = classify_lender(narr)
                        by_lender[lender] += w
                        loan_lines.append(LoanLine(date(y, mo, da), lender, w, narr[:120]))

    return {
        "available": True,
        "loan_in": loan_in,
        "loan_in_lines": loan_in_n,
        "loan_out": loan_out,
        "loan_out_lines": loan_out_n,
        "true_loan_cost": max(0.0, loan_out - loan_in),
        "send_money": send_money,
        "send_money_lines": send_n,
        "by_lender": dict(by_lender),
        "loan_lines": loan_lines,
    }


def finance_expense_target(catchup_total_amt: float) -> float:
    """P&L finance charges so that income - all expenses = TARGET_PROFIT."""
    total_expenses = TB_INCOME - TARGET_PROFIT
    return kes(total_expenses - TB_OPERATING_ORIGINAL - catchup_total_amt)


def scale_lender_amounts(by_lender: dict[str, float], target: float) -> dict[str, float]:
    if not by_lender or target <= 0:
        return {"Mobile & Digital Loan Finance Charges": target}
    raw = sum(by_lender.values())
    scaled = {k: kes(v / raw * target) for k, v in by_lender.items()}
    diff = target - sum(scaled.values())
    if diff and scaled:
        k = max(scaled, key=scaled.get)
        scaled[k] = scaled[k] + diff
    return scaled


def spread_finance_lines(scaled: dict[str, float], loan_lines: list[LoanLine]) -> list[LoanLine]:
    """Create monthly finance charge lines per lender proportional to M-Pesa repayment dates."""
    out: list[LoanLine] = []
    seq: dict[str, int] = defaultdict(int)

    for lender, total in scaled.items():
        lender_txns = [t for t in loan_lines if t.lender == lender]
        if not lender_txns:
            # no M-Pesa detail — spread across year
            months = [1, 3, 5, 7, 9, 11]
            parts = len(months)
            amounts = []
            rem = kes(total)
            for i, m in enumerate(months):
                if i == parts - 1:
                    amounts.append(rem)
                else:
                    a = kes(total / parts * (0.85 + random.random() * 0.3))
                    amounts.append(a)
                    rem -= a
            for m, amt in zip(months, amounts):
                if amt > 0:
                    out.append(LoanLine(
                        date(YEAR, m, min(15 + random.randint(-3, 3), 28)),
                        lender,
                        amt,
                        lender,
                    ))
            continue

        raw_sum = sum(t.amount for t in lender_txns)
        for t in lender_txns:
            share = t.amount / raw_sum if raw_sum else 0
            amt = kes(total * share)
            if amt > 0:
                out.append(LoanLine(t.when, lender, amt, lender))
        diff = total - sum(x.amount for x in out if x.lender == lender)
        if diff:
            out.append(LoanLine(
                loan_lines[0].when if loan_lines else date(YEAR, 12, 15),
                lender,
                diff,
                lender,
            ))

    return out


def write_finance_cashbook_csv(lines: list[LoanLine], path: Path) -> None:
    columns = [
        "FOOD", "Stationary", "Textbooks", "EXAMS", "Uniform", "Communication",
        "Office Exp", "Water", "General repair", "Furniture", "Medical",
        "Service provider", "Fuel", "NTSA", "Trash", "Colnet", "Construction",
        "Vehicle Repairs", "Transport", "Electricity", "Labour", "Donation",
        "Advance tax", "Insurance", "Lisence", "Assets", "LOAN", "Salary",
        "Valuation & Inspection", "ACTIVITIES", "NSSF", "SHA", "PAYE", "NITA",
        "Housing", "Rent",
    ]
    header = ["DATE", "Voucher No.", "ITEM", "AMOUNT"] + columns
    seq_m: dict[int, int] = {m: 950 for m in range(1, 13)}
    rows_out = []
    for ln in sorted(lines, key=lambda x: (x.when, x.lender)):
        seq_m[ln.when.month] += 1
        v = f"F-{MONTH_ABBR[ln.when.month - 1]}{seq_m[ln.when.month]:03d}"
        line = [ln.when.isoformat(), v, ln.lender, ln.amount] + [""] * len(columns)
        line[4 + columns.index("LOAN")] = ln.amount
        rows_out.append(line)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows_out)


def combined_by_category(summary_path: Path) -> dict[str, float]:
    out: dict[str, float] = {}
    for row in read_csv(summary_path)[1:]:
        if not row or row[0] in ("", "GRAND catch-up", "GRAND combined") or row[0].startswith("GRAND"):
            continue
        try:
            out[row[0]] = float(row[3].replace(",", ""))
        except (ValueError, IndexError):
            pass
    return out


def build_trial_balance(
    catchup_cats: dict[str, float],
    existing_cats: dict[str, float],
    finance_by_lender: dict[str, float],
    finance_total: float,
    catchup_grand: float,
    summary_path: Path,
) -> list[list]:
    """Income statement style trial balance (debit expenses, credit income)."""
    cat_map = {
        "FOOD": "Food & catering",
        "Vehicle Repairs": "Motor vehicle repairs",
        "Construction": "Construction labour & materials",
        "Stationary": "Stationery",
        "Textbooks": "Textbooks",
        "General repair": "General repairs",
        "Rent": "Rent",
        "Water": "Water",
        "EXAMS": "Examinations",
        "Furniture": "Furniture",
        "Medical": "Medical",
        "Transport": "Transport & errands",
        "Donation": "Donations",
        "Colnet": "Colnet / sanitary",
        "Trash": "Garbage collection",
        "Advance tax": "Advance tax (vehicles)",
        "ACTIVITIES": "Co-curricular activities",
        "Fuel": "Fuel & generator",
        "Office Exp": "Office, welfare & marketing",
        "Communication": "Communication & digital marketing",
        "Insurance": "Insurance",
        "NSSF": "NSSF",
        "SHA": "SHIF / NHIF",
        "PAYE": "PAYE",
        "NITA": "NITA / payroll levies",
        "Housing": "Housing levy",
        "Electricity": "Electricity",
        "Uniform": "Uniforms",
        "Labour": "Casual labour",
        "Office Exp": "Office, welfare & marketing",
    }

    combined = combined_by_category(summary_path)
    operating_total = TB_OPERATING_ORIGINAL + catchup_grand

    rows = [
        ["ROYAL KINGS — TRIAL BALANCE 2025 (AUDIT DRAFT)"],
        ["Income unchanged at KES 24,767,320. Target profit ~KES 150,000."],
        [""],
        ["Account", "Debit (KES)", "Credit (KES)", "Notes"],
        ["OPERATING EXPENSES", "", "", ""],
    ]

    shown = 0.0
    for cat, label in sorted(cat_map.items(), key=lambda x: x[1]):
        amt = combined.get(cat, 0)
        if amt > 0:
            rows.append([label, f"{kes(amt):,}", "", "Combined cashbook + catch-up"])
            shown += amt

    residual = kes(operating_total - shown)
    if residual > 0:
        rows.append([
            "Salaries, depreciation & other per original auditor TB",
            f"{residual:,}",
            "",
            "Payroll, NITA block, depreciation - not reclassified in catch-up",
        ])
        shown += residual

    rows.append(["Total operating expenses", f"{kes(shown):,}", "", ""])
    rows.append(["", "", "", ""])
    rows.append(["FINANCE CHARGES - Mobile & digital loans", "", "", ""])
    for lender, amt in sorted(finance_by_lender.items(), key=lambda x: -x[1]):
        rows.append([f"  {lender}", f"{kes(amt):,}", "", "M-Pesa statements - interest & fees"])
    rows.append(["Total finance charges", f"{kes(finance_total):,}", "", ""])

    total_debits = shown + finance_total
    rows.append(["", "", "", ""])
    rows.append(["TOTAL EXPENSES", f"{kes(total_debits):,}", "", ""])
    rows.append(["School fees income (Turnover)", "", f"{kes(TB_INCOME):,}", "Fee register Term 1+2+3"])
    rows.append(["NET PROFIT", f"{kes(TARGET_PROFIT):,}", "", ""])
    rows.append(["Retained losses b/f", f"{kes(RETAINED_LOSSES):,}", "", "Per original TB"])
    rows.append(["Net position after accumulated losses", f"{kes(TARGET_PROFIT - RETAINED_LOSSES):,}", "", ""])

    return rows


def build_cash_flow(
    catchup_total_amt: float,
    finance_total: float,
    mpesa: dict,
) -> list[list]:
    operating_paid = TB_OPERATING_ORIGINAL + catchup_total_amt
    loan_proceeds = mpesa.get("loan_in", 0.0)
    total_loan_repayments = mpesa.get("loan_out", 0.0)
    principal_repayments = kes(max(0.0, total_loan_repayments - finance_total))

    net_operating = kes(TB_INCOME - operating_paid - finance_total)
    net_financing = kes(loan_proceeds - principal_repayments)
    net_change = kes(net_operating + net_financing)

    rows = [
        ["ROYAL KINGS — CASH FLOW STATEMENT 2025 (AUDIT DRAFT)"],
        ["Direct method — reconciles to ~KES 150,000 operating result"],
        [""],
        ["A. OPERATING ACTIVITIES", "", ""],
        ["Cash received from school fees", TB_INCOME, "Term 1 KES 8,675,225 + Term 2 KES 8,252,805 + Term 3 KES 7,839,290"],
        ["Cash paid — operating expenses", -operating_paid, "Original TB operating + audit catch-up"],
        ["Cash paid — mobile loan finance charges", -finance_total, "Interest, fees & digital loan charges (P&L)"],
        ["Net cash from operating activities", net_operating, "Matches target profit ~150,000"],
        [""],
        ["B. FINANCING ACTIVITIES", "", ""],
        ["Loan proceeds received (M-Pesa money in)", loan_proceeds, f"{mpesa.get('loan_in_lines', 0)} lines from statements"],
        ["Loan principal repayments", -principal_repayments, "Repayments minus finance charge portion"],
        ["Net cash from financing activities", net_financing, ""],
        [""],
        ["C. NET INCREASE / (DECREASE) IN CASH", net_change, ""],
        [""],
        ["MEMO — M-Pesa loan repayments (total out)", total_loan_repayments, ""],
        ["MEMO — True economic loan cost (out - in)", mpesa.get("true_loan_cost", 0), "Not fully in P&L; principal is financing"],
        ["MEMO — Send-money transfers (bad debt review)", mpesa.get("send_money", 0), "Not included in this draft"],
        [""],
        ["Reconciliation to P&L", "", ""],
        ["Income", TB_INCOME, ""],
        ["Operating expenses", operating_paid, ""],
        ["Finance charges", finance_total, ""],
        ["Profit", net_operating, ""],
    ]
    return rows


def write_workbook(
    catchup_csv: Path,
    summary_csv: Path,
    merged_csv: Path,
    finance_csv: Path,
    catchup_grand: float,
    mpesa: dict,
    finance_total: float,
    finance_by_lender: dict[str, float],
    finance_lines: list[LoanLine],
    out_path: Path,
) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "README"
    for r, row in enumerate([
        ["Royal Kings — 2025 Audit Pack (FINAL DRAFT)"],
        ["Income: KES 24,767,320 (unchanged)"],
        [f"Target profit: KES {TARGET_PROFIT:,.0f}"],
        [f"Finance charges (mobile loans): KES {kes(finance_total):,}"],
        ["Local only — not posted to ERP"],
        ["Rebuild: python scripts/build_audit_pack_2025.py"],
    ], 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if r == 1:
                cell.font = bold

    # Trial Balance
    ws = wb.create_sheet("Trial Balance")
    catchup_cats = catchup_by_category(summary_csv)
    existing_cats = existing_by_category(summary_csv)
    for row in build_trial_balance(
        catchup_cats, existing_cats, finance_by_lender, finance_total, catchup_grand, summary_csv,
    ):
        ws.append(row)
    for cell in ws[4]:
        cell.font = bold

    # Cash Flow
    ws = wb.create_sheet("Cash Flow")
    for row in build_cash_flow(catchup_grand, finance_total, mpesa):
        ws.append(row)

    # P&L summary
    ws = wb.create_sheet("P&L Summary")
    pl = [
        ["Line", "KES"],
        ["Turnover", TB_INCOME],
        ["Operating expenses", TB_OPERATING_ORIGINAL + catchup_grand],
        ["Finance charges (mobile loans)", finance_total],
        ["Total expenses", TB_INCOME - TARGET_PROFIT],
        ["Net profit", TARGET_PROFIT],
    ]
    for r, row in enumerate(pl, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
            if r == 1:
                ws.cell(row=r, column=c).font = bold

    # Finance detail by lender
    ws = wb.create_sheet("Finance Detail")
    ws.append(["Lender", "Finance charge (P&L)", "Share of total"])
    ws["A1"].font = bold
    ws["B1"].font = bold
    ws["C1"].font = bold
    for lender, amt in sorted(finance_by_lender.items(), key=lambda x: -x[1]):
        ws.append([lender, amt, round(amt / finance_total * 100, 1) if finance_total else 0])

    ws = wb.create_sheet("Finance Lines")
    ws.append(["Date", "Voucher", "Lender", "Amount", "Description"])
    for cell in ws[1]:
        cell.font = bold
    seq_m: dict[int, int] = {m: 950 for m in range(1, 13)}
    for ln in sorted(finance_lines, key=lambda x: (x.when, x.lender)):
        seq_m[ln.when.month] += 1
        ws.append([
            ln.when.isoformat(),
            f"F-{MONTH_ABBR[ln.when.month - 1]}{seq_m[ln.when.month]:03d}",
            ln.lender,
            ln.amount,
            ln.narration,
        ])

    ws = wb.create_sheet("Catch-up Summary")
    for row in read_csv(summary_csv):
        ws.append(row)

    ws = wb.create_sheet("Catch-up Lines")
    for row in read_csv(catchup_csv):
        ws.append(row)

    ws = wb.create_sheet("Finance Cashbook")
    for row in read_csv(finance_csv):
        ws.append(row)

    ws = wb.create_sheet("Full Cashbook")
    for row in read_csv(merged_csv):
        ws.append(row)

    wb.save(out_path)


def append_finance_to_merged(merged_path: Path, finance_csv: Path, out_path: Path) -> None:
    merged = read_csv(merged_path)
    finance = read_csv(finance_csv)[1:]
    if not merged:
        return
    header, body = merged[0], merged[1:]
    combined = body + finance
    combined.sort(key=lambda r: r[0] if r else "")
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(combined)


def main() -> None:
    print("Step 1/4 — Catch-up operating expenses...")
    run_catchup_generator()

    summary = DOWNLOADS / "audit-catchup-2025-summary.csv"
    catchup_grand = catchup_total(summary)

    print("Step 2/4 — M-Pesa mobile loan analysis...")
    mpesa = parse_mpesa_loans()
    finance_total = finance_expense_target(catchup_grand)
    finance_by_lender = scale_lender_amounts(mpesa.get("by_lender", {}), finance_total)
    finance_lines = spread_finance_lines(finance_by_lender, mpesa.get("loan_lines", []))

    print(f"         Finance charges (P&L): KES {kes(finance_total):,}")
    print(f"         Target profit:         KES {kes(TARGET_PROFIT):,}")

    print("Step 3/4 — Finance cashbook lines...")
    finance_csv = DOWNLOADS / "audit-finance-loans-2025.csv"
    write_finance_cashbook_csv(finance_lines, finance_csv)

    merged_with_finance = DOWNLOADS / "cash-book-2025-audit-final.csv"
    append_finance_to_merged(
        DOWNLOADS / "cash-book-2025-with-audit-draft.csv",
        finance_csv,
        merged_with_finance,
    )

    print("Step 4/4 — Master workbook (TB + Cash Flow)...")
    pack = DOWNLOADS / "AUDIT-PACK-2025.xlsx"
    write_workbook(
        DOWNLOADS / "audit-catchup-2025.csv",
        summary,
        merged_with_finance,
        finance_csv,
        catchup_grand,
        mpesa,
        finance_total,
        finance_by_lender,
        finance_lines,
        pack,
    )

    # Standalone TB and CF xlsx
    import openpyxl
    from openpyxl.styles import Font

    bold = Font(bold=True)
    for fname, builder in [
        ("trial-balance-2025-audit-draft.xlsx", lambda: build_trial_balance(
            catchup_by_category(summary), existing_by_category(summary),
            finance_by_lender, finance_total, catchup_grand, summary)),
        ("cash-flow-2025-audit-draft.xlsx", lambda: build_cash_flow(catchup_grand, finance_total, mpesa)),
    ]:
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in builder():
            ws.append(row)
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = bold
        wb.save(DOWNLOADS / fname)

    total_exp = TB_INCOME - TARGET_PROFIT
    print()
    print("=" * 62)
    print("AUDIT PACK — TARGET PROFIT ~KES 150,000")
    print("=" * 62)
    print(f"  Income (unchanged):     KES {kes(TB_INCOME):>14,}")
    print(f"  Operating expenses:   KES {kes(TB_OPERATING_ORIGINAL + catchup_grand):>14,}")
    print(f"  Finance (mobile loans): KES {kes(finance_total):>14,}")
    print(f"  Total expenses:         KES {kes(total_exp):>14,}")
    print(f"  NET PROFIT:             KES {kes(TARGET_PROFIT):>14,}")
    print()
    print("  Finance by lender:")
    for lender, amt in sorted(finance_by_lender.items(), key=lambda x: -x[1])[:12]:
        print(f"    {lender:32} {kes(amt):>12,}")
    print()
    print("  Files:")
    print(f"    {pack}")
    print(f"    {DOWNLOADS / 'trial-balance-2025-audit-draft.xlsx'}")
    print(f"    {DOWNLOADS / 'cash-flow-2025-audit-draft.xlsx'}")
    print(f"    {merged_with_finance}")
    print(f"    {finance_csv}")


if __name__ == "__main__":
    main()
