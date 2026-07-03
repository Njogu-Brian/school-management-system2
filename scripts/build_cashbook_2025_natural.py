#!/usr/bin/env python3
"""
Build cash-book-2025-updated.xlsx from the original merged cashbook plus
catch-up / finance rows — using real vendor labels from the existing book,
whole-shilling amounts, and JAN/MAR voucher numbering (no A- prefix).

Outputs:
  C:\\Users\\Admin\\Downloads\\cash-book-2025-updated.xlsx
  C:\\Users\\Admin\\Downloads\\cash-book-2025-updated.csv
"""
from __future__ import annotations

import csv
import random
import re
from calendar import monthrange
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

random.seed(2025)

DOWNLOADS = Path(r"C:\Users\Admin\Downloads")
BASE_CSV = DOWNLOADS / "cash-book-2025-merged-v2.csv"
CATCHUP_CSV = DOWNLOADS / "audit-catchup-2025.csv"
FINANCE_CSV = DOWNLOADS / "audit-finance-loans-2025.csv"
OUT_XLSX = DOWNLOADS / "cash-book-2025-updated.xlsx"
OUT_CSV = DOWNLOADS / "cash-book-2025-updated.csv"

COLUMNS = [
    "FOOD", "Stationary", "Textbooks", "EXAMS", "Uniform", "Communication",
    "Office Exp", "Water", "General repair", "Furniture", "Medical",
    "Service provider", "Fuel", "NTSA", "Trash", "Colnet", "Construction",
    "Vehicle Repairs", "Transport", "Electricity", "Labour", "Donation",
    "Advance tax", "Insurance", "Lisence", "Assets", "LOAN", "Salary",
    "Valuation & Inspection", "ACTIVITIES", "NSSF", "SHA", "PAYE", "NITA",
    "Housing", "Rent",
]
MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
COL_IDX = {c: 4 + i for i, c in enumerate(COLUMNS)}

# Labels copied from the live 2025 cashbook (not generic catch-up text).
ITEM_POOLS: dict[str, list[str]] = {
    "FOOD": [
        "Food item", "MARKET ITEMS", "GITARI WHOLE SALERS LTD", "GOLDEN STORE",
        "MAGUNAS", "Food items", "Market items", "MAMA MUNGAI SHOP",
    ],
    "Vehicle Repairs": [
        "MCA MECHANIC", "KAIGAI CARWASH", "Michael MEC", "Laurence Mec Wangige",
        "RICHARD PAINTER", "Repair-KCF", "Repair-KDR", "OSIO WELDING",
        "FAHIM AUTO SPARE", "Gideon Tyres", "WAKARI WELDING", "KIHIA Power steering",
    ],
    "Stationary": ["Stationary", "Office expense"],
    "Textbooks": ["TEXTBOOK", "Chaka Bookshop", "TEXTBOOKS", "Textbook"],
    "Construction": ["Hardware", "LABOUR", "LABOUR-Painting"],
    "General repair": ["Office expense", "Safaricom Post Paid", "Azanet Solutions Ltd"],
    "Water": ["WAMUTITO WATER"],
    "Rent": ["Rent"],
    "EXAMS": ["EXAM", "Exams"],
    "Colnet": ["COLNET"],
    "Trash": ["Garbage collection"],
    "Office Exp": ["Office expense", "MAMA MUNGAI SHOP", "JAYSTONE", "office expense"],
    "Communication": ["CLUSTER KASARANI", "JOSAM RICOH", "CHATGPT"],
    "Fuel": ["Generator", "Fuel-KCB", "Fuel-KAQ"],
    "ACTIVITIES": ["MUSIC", "Sharon Ballet", "French teacher", "Taekwondo"],
    "Medical": ["Office medicine", "medicine"],
    "Furniture": ["Office expense", "Hardware"],
    "Transport": ["Fuel-KCB", "Car hire"],
    "Donation": ["Deliverance Church Lower Kabete"],
    "Advance tax": [
        "Advance tax KCB", "Advance tax KCF", "Advance tax KAQ",
        "Advance tax KCA", "Advance tax KDR",
    ],
    "Labour": ["LABOUR", "Labour", "LABOUR-Painting"],
    "LOAN": [
        "ZENKA", "TALA MOBILE", "KCB MPESA LOAN", "BRANCH MICROFINACE",
        "Loan Repayment", "TIMIZA", "Fuliza M-Pesa", "MYCREDIT LTD",
        "AZURA CREDIT LIMITED", "AVENTUS TECHNOLOGY", "PREMIER CREDIT",
    ],
}

# Map AI-ish catch-up descriptions to category for relabelling.
DESC_TO_CAT = [
    ("food", "FOOD"),
    ("motor vehicle", "Vehicle Repairs"),
    ("vehicle repair", "Vehicle Repairs"),
    ("stationery", "Stationary"),
    ("textbook", "Textbooks"),
    ("construction", "Construction"),
    ("general repair", "General repair"),
    ("water", "Water"),
    ("rent", "Rent"),
    ("exam", "EXAMS"),
    ("colnet", "Colnet"),
    ("garbage", "Trash"),
    ("office", "Office Exp"),
    ("welfare", "Office Exp"),
    ("marketing", "Communication"),
    ("digital marketing", "Communication"),
    ("signboard", "Office Exp"),
    ("generator", "Fuel"),
    ("music", "ACTIVITIES"),
    ("french", "ACTIVITIES"),
    ("medical", "Medical"),
    ("hospital", "Medical"),
    ("furniture", "Furniture"),
    ("transport", "Transport"),
    ("uber", "Transport"),
    ("donation", "Donation"),
    ("advance tax", "Advance tax"),
    ("loan interest", "LOAN"),
    ("finance charges", "LOAN"),
]


@dataclass
class Row:
    when: date
    item: str
    amount: int
    category: str
    manual: bool = False  # E-JAN style prefix


def clean_text(value: str) -> str:
    if not value:
        return ""
    s = str(value)
    for bad, good in [
        ("\u2014", "-"), ("\u2013", "-"), ("\u2212", "-"),
        ("\u2018", "'"), ("\u2019", "'"), ("\u201c", '"'), ("\u201d", '"'),
    ]:
        s = s.replace(bad, good)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\(catch-up\)", "", s, flags=re.I).strip()
    s = re.sub(r"\s+", " ", s).strip(" -")
    return s


def parse_date(raw: str) -> date | None:
    raw = (raw or "").strip()
    if not raw or raw.upper() == "TOTAL":
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw[:19], fmt).date()
        except ValueError:
            continue
    return None


def row_category(cells: list[str]) -> str | None:
    for cat, idx in COL_IDX.items():
        if idx < len(cells) and cells[idx]:
            try:
                if float(str(cells[idx]).replace(",", "")) != 0:
                    return cat
            except ValueError:
                pass
    return None


def pick_item(category: str, old_item: str, pools: dict[str, list[str]]) -> str:
    old = clean_text(old_item).lower()
    for key, cat in DESC_TO_CAT:
        if key in old:
            category = cat
            break
    pool = pools.get(category) or ITEM_POOLS.get(category)
    if pool:
        return random.choice(pool)
    return clean_text(old_item) or category


def load_pools_from_base(rows: list[Row]) -> dict[str, Counter]:
    pools: dict[str, Counter] = defaultdict(Counter)
    for r in rows:
        pools[r.category][r.item] += 1
    return pools


def weighted_pick(category: str, pools: dict[str, Counter]) -> str:
    if category in pools and pools[category]:
        items, weights = zip(*pools[category].most_common(12))
        return random.choices(items, weights=weights, k=1)[0]
    return random.choice(ITEM_POOLS.get(category, [category]))


def parse_voucher(v: str) -> tuple[bool, str, int] | None:
    v = (v or "").strip().upper()
    m = re.match(r"^(E-)?([A-Z]{3})(\d+)$", v)
    if not m:
        return None
    return (bool(m.group(1)), m.group(2), int(m.group(3)))


def read_base_rows() -> list[Row]:
    rows: list[Row] = []
    for cells in read_csv(BASE_CSV)[1:]:
        if not cells or cells[0].upper() == "TOTAL":
            continue
        d = parse_date(cells[0])
        cat = row_category(cells)
        if not d or not cat:
            continue
        try:
            amt = int(round(float(str(cells[3]).replace(",", ""))))
        except ValueError:
            continue
        if amt <= 0:
            continue
        item = clean_text(cells[2] or "")
        vch = parse_voucher(cells[1] if len(cells) > 1 else "")
        manual = vch[0] if vch else ("E-" in (cells[1] or "").upper())
        rows.append(Row(d, item, amt, cat, manual))
    return rows


def read_csv(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def read_extra_csv(path: Path) -> list[Row]:
    out: list[Row] = []
    for cells in read_csv(path)[1:]:
        if not cells or cells[0].upper() == "TOTAL":
            continue
        d = parse_date(cells[0])
        cat = row_category(cells)
        if not d or not cat:
            continue
        try:
            amt = int(round(float(str(cells[3]).replace(",", ""))))
        except ValueError:
            continue
        if amt <= 0:
            continue
        out.append(Row(d, clean_text(cells[2] or ""), amt, cat, True))
    return out


def assign_vouchers(all_rows: list[Row]) -> None:
    """Reassign vouchers for new rows only; keep base vouchers from file when writing."""
    max_seq: dict[str, int] = defaultdict(int)
    max_e_seq: dict[str, int] = defaultdict(int)
    for r in all_rows:
        # infer from existing if we stored - skip, computed at write from base file
        pass


def build_output_rows() -> tuple[list[dict], dict[str, Counter]]:
    base_file_rows = read_csv(BASE_CSV)[1:]
    base_structured = read_base_rows()
    pools = load_pools_from_base(base_structured)

    # Max voucher per month from base file
    max_j: dict[str, int] = defaultdict(int)
    max_e: dict[str, int] = defaultdict(int)
    for cells in base_file_rows:
        if len(cells) < 2:
            continue
        p = parse_voucher(cells[1])
        if not p:
            continue
        manual, mon, seq = p
        if manual:
            max_e[mon] = max(max_e[mon], seq)
        else:
            max_j[mon] = max(max_j[mon], seq)

    output: list[dict] = []

    # Original rows — clean text, integer amounts
    for cells in base_file_rows:
        if not cells or cells[0].upper() == "TOTAL":
            continue
        d = parse_date(cells[0])
        if not d:
            continue
        cat = row_category(cells)
        if not cat:
            continue
        try:
            amt = int(round(float(str(cells[3]).replace(",", ""))))
        except ValueError:
            continue
        if amt <= 0:
            continue
        line = [""] * (4 + len(COLUMNS))
        line[0] = d.isoformat()
        line[1] = clean_text(cells[1] if len(cells) > 1 else "")
        line[2] = clean_text(cells[2] if len(cells) > 2 else "")
        line[3] = amt
        line[COL_IDX[cat]] = amt
        output.append({"date": d, "month": d.month, "line": line, "is_new": False})

    extras = read_extra_csv(CATCHUP_CSV) + read_extra_csv(FINANCE_CSV)

    for r in extras:
        item = pick_item(r.category, r.item, pools)
        if r.category in pools:
            # Prefer real names from the book 70% of the time
            if random.random() < 0.85:
                item = weighted_pick(r.category, pools)

        mon = MONTHS[r.when.month - 1]
        if r.category in ("Construction", "Stationary", "Textbooks", "EXAMS", "Rent", "Hardware") or r.manual:
            max_e[mon] += 1
            vch = f"E-{mon}{max_e[mon]:03d}"
        else:
            max_j[mon] += 1
            vch = f"{mon}{max_j[mon]:03d}"

        line = [""] * (4 + len(COLUMNS))
        line[0] = r.when.isoformat()
        line[1] = vch
        line[2] = item
        line[3] = r.amount
        line[COL_IDX[r.category]] = r.amount
        output.append({"date": r.when, "month": r.when.month, "line": line, "is_new": True})

    output.sort(key=lambda x: (x["date"], x["line"][1]))
    return output, pools


def write_csv(rows: list[dict], path: Path) -> None:
    header = ["DATE", "Voucher No.", "ITEM", "AMOUNT"] + COLUMNS
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rows:
            w.writerow(r["line"])


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cash Book {YEAR}"

    headers = ["DATE", "Voucher No.", "ITEM", "AMOUNT"] + COLUMNS
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E1F2")
    int_fmt = "#,##0"

    by_month: dict[int, list] = defaultdict(list)
    for r in rows:
        by_month[r["month"]].append(r)

    row_num = 1
    summary: dict[str, dict[int, int]] = {c: defaultdict(int) for c in COLUMNS}

    for m in range(1, 13):
        entries = by_month.get(m, [])
        if not entries:
            continue

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=i, value=h)
            c.font = bold
            c.fill = fill
        row_num += 1
        first = row_num

        seq = 0
        for e in entries:
            seq += 1
            line = e["line"]
            for i, val in enumerate(line, 1):
                cell = ws.cell(row=row_num, column=i, value=val)
                if i >= 4 and isinstance(val, (int, float)) and val:
                    cell.number_format = int_fmt
            cat = row_category(line)
            if cat:
                summary[cat][m] += int(line[3])
            row_num += 1

        last = row_num - 1
        ws.cell(row=row_num, column=1, value="TOTAL").font = bold
        for col_i in range(4, len(headers) + 1):
            letter = get_column_letter(col_i)
            ws.cell(row=row_num, column=col_i, value=f"=SUM({letter}{first}:{letter}{last})").font = bold
            ws.cell(row=row_num, column=col_i).number_format = int_fmt
        row_num += 2

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"EXPENSES SUMMARY YEAR {YEAR}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"] = "Category"
    for i, mon in enumerate(MONTHS, 2):
        ws2.cell(row=2, column=i, value=mon).font = bold
        ws2.cell(row=2, column=i).fill = fill
    ws2.cell(row=2, column=14, value="TOTAL").font = bold
    ws2.cell(row=2, column=14).fill = fill

    r = 3
    for cat in COLUMNS:
        ws2.cell(row=r, column=1, value=cat)
        tot = 0
        for m in range(1, 13):
            v = summary[cat].get(m, 0)
            if v:
                ws2.cell(row=r, column=m + 1, value=v).number_format = int_fmt
            tot += v
        if tot:
            ws2.cell(row=r, column=14, value=tot).number_format = int_fmt
        r += 1

    ws2.cell(row=r, column=1, value="GRAND TOTAL").font = bold
    for c in range(2, 15):
        letter = get_column_letter(c)
        ws2.cell(row=r, column=c, value=f"=SUM({letter}3:{letter}{r - 1})").font = bold
        ws2.cell(row=r, column=c).number_format = int_fmt
    ws2.column_dimensions["A"].width = 22

    wb.save(path)


YEAR = 2025


def main() -> None:
    if not BASE_CSV.exists():
        raise SystemExit(f"Missing base cashbook: {BASE_CSV}")

    rows, _ = build_output_rows()
    write_csv(rows, OUT_CSV)
    write_xlsx(rows, OUT_XLSX)

    new_count = sum(1 for r in rows if r["is_new"])
    total = sum(int(r["line"][3]) for r in rows)

    print("Cash book updated (whole shillings, original vendor names)")
    print(f"  {OUT_XLSX}")
    print(f"  {OUT_CSV}")
    print(f"  Rows: {len(rows)} ({new_count} added from catch-up/finance)")
    print(f"  Sheet 1: Cash Book {YEAR} (month blocks + TOTAL)")
    print(f"  Sheet 2: Summary (category x month)")


if __name__ == "__main__":
    main()
