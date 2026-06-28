"""
Fill EXPENSES-2025.xlsx with expenses that exist in the accounting system but are
not yet recorded in the spreadsheet.

- Reads a mysqldump .sql.gz backup (expenses, expense_lines, expense_categories, vendors).
- Reads the workbook and figures out each month block (header row .. TOTAL row).
- For every system expense line, maps its DB category to the matching spreadsheet
  column and inserts a new row into that month's block IF it is not already there.
- "Already there" = a row in the same category column whose amount matches within
  KES 1 and whose date is within +/-10 days (so we don't double-enter).
- Writes a NEW file; the original workbook is never modified.

Usage:
  python scripts/fill_expenses_workbook.py <dump.sql.gz> <input.xlsx> <output.xlsx> [--apply]

Without --apply it only prints the analysis (dry run).
"""
import sys, gzip, re, datetime, collections
import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- SQL dump parsing
def grab_values(path, table):
    pat = re.compile(r"INSERT INTO `%s` VALUES " % re.escape(table))
    out = []
    with gzip.open(path, 'rt', encoding='utf-8', errors='replace') as f:
        for line in f:
            m = pat.search(line)
            if m:
                out.append(line[m.end():].rstrip().rstrip(';'))
    return out

def split_tuples(blob):
    depth = 0; cur = []; in_str = False; esc = False
    for ch in blob:
        if in_str:
            cur.append(ch)
            if esc: esc = False
            elif ch == '\\': esc = True
            elif ch == "'": in_str = False
        else:
            if ch == "'":
                in_str = True; cur.append(ch)
            elif ch == '(':
                if depth == 0: cur = []
                else: cur.append(ch)
                depth += 1
            elif ch == ')':
                depth -= 1
                if depth == 0: yield ''.join(cur)
                else: cur.append(ch)
            elif depth > 0:
                cur.append(ch)

def split_fields(tup):
    fields = []; cur = []; in_str = False; esc = False; was_str = False
    for ch in tup:
        if in_str:
            if esc:
                cur.append(ch); esc = False
            elif ch == '\\': esc = True
            elif ch == "'": in_str = False
            else: cur.append(ch)
        else:
            if ch == "'":
                in_str = True; was_str = True
            elif ch == ',':
                val = ''.join(cur)
                fields.append(None if (not was_str and val.strip() == 'NULL') else val.strip())
                cur = []; was_str = False
            else:
                cur.append(ch)
    val = ''.join(cur)
    fields.append(None if (not was_str and val.strip() == 'NULL') else val.strip())
    return fields

def parse_table(path, table):
    rows = []
    for blob in grab_values(path, table):
        for tup in split_tuples(blob):
            rows.append(split_fields(tup))
    return rows

# ---------------------------------------------------------------- category -> column map
# Spreadsheet columns (exact header text in Sheet1, row 1, from col E onward)
# DB category CODE -> spreadsheet column header
CODE_TO_COLUMN = {
    'FUEL': 'Fuel',
    'VEH-REPAIRS': 'Vehicle Repairs', 'VEH-SERVICE': 'Vehicle Repairs',
    'SPEED-GOVERNOR': 'Vehicle Repairs', 'VEH-TRACKING': 'Vehicle Repairs',
    'VEH-INSURANCE': 'Insuarance',
    'VEH-INSPECTION': 'Valuation & Inspection', 'VEH-VALUATION': 'Valuation & Inspection',
    'LAND-VALUATION': 'Valuation & Inspection',
    'VEH-LOGBOOK': 'NTSA', 'NTSA': 'NTSA',
    'CAR-HIRE': 'transport', 'TRANSPORT': 'transport',
    'VEH-ADVANCE-TAX': 'Advance tax', 'ADVANCE-TAX': 'Advance tax',
    'VEH-PURCHASE': 'Assests', 'ASSETS': 'Assests', 'ADMIN-COMPUTER-EQUIPMENT': 'Assests',
    'SALARIES': 'Salary', 'WAGES': 'Salary', 'STAFF': 'Salary',
    'MEDICAL': 'Medical',
    'PAYE': 'PAYE', 'NSSF': 'NSSF', 'NHIF': 'SHA', 'HOUSING': 'Housing', 'NITA': 'NITA',
    'LOANS': 'LOAN', 'LOAN-EQUITY-8659': 'LOAN', 'LOAN-EQUITY-2564': 'LOAN',
    'LOAN-EQUITY-986': 'LOAN', 'LOAN-EQUITY-7419': 'LOAN', 'LOAN-IM-BANK': 'LOAN',
    'LOAN-FAMILY-BANK': 'LOAN', 'LOAN-JACKFRUIT': 'LOAN', 'LOAN-ED-PARTNERS': 'LOAN',
    'LOANS-TCL-CREDIT': 'LOAN',
    'ELECTRICITY': 'Electricity', 'WATER': 'Water',
    'INTERNET': 'Communication', 'WIFI': 'Communication', 'COMMUNICATION': 'Communication',
    'TRASH': 'Trash', 'SANITARY': 'Colnet',
    'OFFICE': 'Office Exp', 'ADMIN': 'Office Exp', 'AUDIT-FEE': 'Office Exp',
    'STATIONERY': 'Stationary', 'LICENSE': 'Lisence',
    'RENT': 'Rent', 'DONATION': 'Donation', 'FURNITURE': 'Furniture',
    'TEXTBOOKS': 'Textbooks', 'EXAM': 'EXAMS', 'UNIFORM': 'Uniform',
    'CATERING': 'FOOD', 'FOOD': 'FOOD',
    'ACTIVITIES': 'ACTIVITIES ', 'ACT-BALLET': 'ACTIVITIES ', 'ACT-SKATING': 'ACTIVITIES ',
    'ACT-TAEKWONDO': 'ACTIVITIES ', 'ACT-MUSIC': 'ACTIVITIES ', 'ACT-FRENCH': 'ACTIVITIES ',
    'ACTIVITIES-YORGHUT': 'ACTIVITIES ', 'ACTIVITIES-GRADUATION': 'ACTIVITIES ',
    'SCHOOL-TRIPS': 'ACTIVITIES ',
    'CONSTRUCTION': 'Construction', 'BUILDINGS': 'Construction',
    'LABOUR-CONSTRUCTION': 'Labour', 'GENERAL-REPAIRS': 'General repair',
    'OTHER': 'Office Exp', 'MISC': 'Office Exp', 'OTHER-AI-TOOLS': 'Office Exp',
    'GENERATOR': 'Fuel',
    # TXN_COST (bank/M-Pesa charges) has no column in the sheet -> skipped on purpose.
}
SKIP_CODES = {'TXN_COST'}
MONTHS = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']

def main():
    if len(sys.argv) < 4:
        print(__doc__); sys.exit(1)
    dump, xlsx_in, xlsx_out = sys.argv[1], sys.argv[2], sys.argv[3]
    apply = '--apply' in sys.argv[4:]

    # ---- DB
    cats = {r[0]: {'code': r[1], 'name': r[2], 'parent': r[3]} for r in parse_table(dump, 'expense_categories')}
    vendors = {r[0]: r[1] for r in parse_table(dump, 'vendors')}
    exp_rows = parse_table(dump, 'expenses')
    expenses = {}
    for r in exp_rows:
        expenses[r[0]] = {
            'no': r[1], 'vendor_id': r[3], 'date': r[5], 'total': r[10],
            'status': r[11], 'notes': r[12], 'deleted_at': r[18] if len(r) > 18 else None,
        }
    lines = parse_table(dump, 'expense_lines')

    def col_for_cat(cid):
        seen = set()
        while cid is not None and cid in cats and cid not in seen:
            seen.add(cid)
            code = cats[cid]['code']
            if code in SKIP_CODES:
                return None
            if code in CODE_TO_COLUMN:
                return CODE_TO_COLUMN[code]
            cid = cats[cid]['parent']
        return None

    # Build the list of system expense "entries" to potentially add
    system_entries = []  # dict: date(datetime), amount(float), column, item, code, exp_no
    unmapped = collections.Counter()
    for ln in lines:
        exp_id = ln[1]; cid = ln[2]; desc = ln[5]; line_total = ln[9]
        exp = expenses.get(exp_id)
        if not exp or exp['deleted_at']:
            continue
        if exp['status'] not in ('submitted', 'approved', 'paid'):
            continue
        code = cats.get(cid, {}).get('code')
        if code in SKIP_CODES:
            continue
        column = col_for_cat(cid)
        if column is None:
            unmapped[code or cid] += 1
            continue
        try:
            amount = float(line_total)
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue
        try:
            d = datetime.datetime.strptime(exp['date'], '%Y-%m-%d')
        except (TypeError, ValueError):
            continue
        if d.year != 2025:
            continue
        item = vendors.get(exp['vendor_id']) or (desc[:40] if desc else (cats.get(cid, {}).get('name') or 'Expense'))
        system_entries.append({'date': d, 'amount': amount, 'column': column,
                               'item': item, 'code': code, 'exp_no': exp['no']})

    # ---- Workbook
    wb = openpyxl.load_workbook(xlsx_in)
    ws = wb['Sheet1']
    headers = [c.value for c in ws[1]]
    col_index = {}  # header text -> column number (1-based)
    for idx, h in enumerate(headers, start=1):
        if h is not None and str(h).strip() != '':
            col_index[str(h).strip()] = idx
    # normalize 'ACTIVITIES ' lookups
    def find_col(name):
        if name in col_index: return col_index[name]
        for k, v in col_index.items():
            if k.strip().lower() == name.strip().lower():
                return v
        return None

    AMOUNT_COL = find_col('AMOUNT')
    DATE_COL = find_col('DATE')
    VOUCHER_COL = find_col('Voucher No.')
    ITEM_COL = find_col('ITEM')

    def prefix_of_voucher(v):
        if not v or not isinstance(v, str):
            return None
        m = re.match(r'\s*([A-Za-z]{2,5})', v)
        return m.group(1).upper() if m else None

    def month_of_prefix(p):
        if not p:
            return None
        p3 = p[:3]
        return MONTHS.index(p3) + 1 if p3 in MONTHS else None

    # Blocks are delimited by repeating 'DATE' header rows; a block runs from its
    # header to just before the next header. A block may (NOV/DEC) end in a TOTAL row.
    header_rows = [r for r in range(1, ws.max_row + 1)
                   if str(ws.cell(r, 1).value).strip().upper() == 'DATE']
    total_rows = {r for r in range(1, ws.max_row + 1)
                  if str(ws.cell(r, 1).value).strip().upper() == 'TOTAL'}

    blocks = []
    existing = []  # (date, column_header, amount, row)
    for i, h in enumerate(header_rows):
        end = (header_rows[i + 1] - 1) if i + 1 < len(header_rows) else ws.max_row
        total_row = next((t for t in sorted(total_rows) if h < t <= end), None)
        last_data = (total_row - 1) if total_row else end
        insert_at = total_row if total_row else end + 1

        months = collections.Counter()
        prefixes = collections.Counter()
        maxnum = 0
        for r in range(h + 1, last_data + 1):
            v = ws.cell(r, VOUCHER_COL).value
            p = prefix_of_voucher(v)
            mo = month_of_prefix(p)
            if mo:
                months[mo] += 1
                prefixes[p] += 1
                mnum = re.search(r'(\d+)\s*$', str(v))
                if mnum:
                    maxnum = max(maxnum, int(mnum.group(1)))
            d = ws.cell(r, DATE_COL).value
            dd = d if isinstance(d, datetime.datetime) else None
            for hname, hidx in col_index.items():
                if hidx <= AMOUNT_COL:
                    continue
                val = ws.cell(r, hidx).value
                if isinstance(val, (int, float)) and val:
                    existing.append((dd, hname, float(val), r))

        blocks.append({
            'header': h, 'last_data': last_data, 'total': total_row, 'insert_at': insert_at,
            'month': months.most_common(1)[0][0] if months else None,
            'prefix': prefixes.most_common(1)[0][0] if prefixes else None,
            'maxnum': maxnum,
        })

    block_for_month = {b['month']: b for b in blocks if b['month']}

    # Dedup + decide which entries to add
    def is_duplicate(entry):
        for (dd, hname, amount, row) in existing:
            if hname.strip().lower() != entry['column'].strip().lower():
                continue
            if abs(amount - entry['amount']) > 1.0:
                continue
            if dd is None:
                return row  # same column+amount, undated row -> treat as dup
            if abs((dd - entry['date']).days) <= 10:
                return row
        return None

    to_add = collections.defaultdict(list)  # month -> entries
    dup_count = 0; add_count = 0
    for e in system_entries:
        mo = e['date'].month
        dup = is_duplicate(e)
        if dup:
            dup_count += 1
            continue
        to_add[mo].append(e)
        add_count += 1

    # ---- Report
    print('CATEGORY -> COLUMN unmapped (skipped, need review):', dict(unmapped) or 'none')
    print(f'\nSystem expense lines considered: {len(system_entries)}')
    print(f'  already in workbook (skipped): {dup_count}')
    print(f'  NEW, to be added:              {add_count}')
    print('\nTo add per month:')
    for mo in sorted(to_add):
        tot = sum(e["amount"] for e in to_add[mo])
        print(f'  {MONTHS[mo-1]}: {len(to_add[mo])} rows, KES {tot:,.0f}')

    if not apply:
        print('\n(DRY RUN) re-run with --apply to write', xlsx_out)
        return

    # ---- Insert rows (process blocks bottom-up so row indices stay valid)
    flag_col = max(col_index.values()) + 2  # leave a gap, put AUTO marker
    ws.cell(1, flag_col).value = 'AUTO-ADDED'
    months_desc = sorted(to_add, key=lambda mo: block_for_month.get(mo, {}).get('header', 0), reverse=True)
    for mo in months_desc:
        entries = sorted(to_add[mo], key=lambda e: e['date'])
        blk = block_for_month.get(mo)
        if not blk:
            print(f'  ! No block found for {MONTHS[mo-1]}, skipping {len(entries)} rows')
            continue
        insert_at = blk['insert_at']
        prefix = blk['prefix'] or MONTHS[mo - 1]
        ws.insert_rows(insert_at, amount=len(entries))
        num = blk['maxnum']
        for i, e in enumerate(entries):
            r = insert_at + i
            num += 1
            ws.cell(r, DATE_COL).value = e['date']
            ws.cell(r, VOUCHER_COL).value = f"{prefix}{num:03d}"
            ws.cell(r, ITEM_COL).value = e['item']
            ws.cell(r, AMOUNT_COL).value = e['amount']
            cc = find_col(e['column'])
            if cc:
                ws.cell(r, cc).value = e['amount']
            ws.cell(r, flag_col).value = 'AUTO'

    # Final pass: recompute every TOTAL row in FINAL coordinates (sum from the
    # nearest header above it down to the row before the TOTAL).
    final_headers = [r for r in range(1, ws.max_row + 1)
                     if str(ws.cell(r, 1).value).strip().upper() == 'DATE']
    for t in range(1, ws.max_row + 1):
        if str(ws.cell(t, 1).value).strip().upper() != 'TOTAL':
            continue
        hdr_above = max([h for h in final_headers if h < t], default=1)
        first_data = hdr_above + 1
        last_data = t - 1
        if last_data < first_data:
            continue
        for hname, hidx in col_index.items():
            if hidx < AMOUNT_COL:
                continue
            letter = get_column_letter(hidx)
            ws.cell(t, hidx).value = f"=SUM({letter}{first_data}:{letter}{last_data})"

    wb.save(xlsx_out)
    print(f'\nSaved {xlsx_out}. Added {add_count} rows across {len(to_add)} months. Original left untouched.')

if __name__ == '__main__':
    main()
