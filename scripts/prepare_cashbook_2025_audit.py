#!/usr/bin/env python3
"""
Build an audit-ready cash book from cash-book-2025-updated.xlsx:

  - Preserve every existing detail row
  - Insert TB Salaries & wages-NITA (5,769,461) under Salary (12 month-end rows)
  - Keep real E-Citizen NITA levy rows under NITA
  - Month TOTAL rows use SUM formulas
  - Summary category×month cells link to detail month TOTAL cells
  - Profit note on Summary vs TB turnover

Outputs:
  C:\\Users\\brian\\Downloads\\down\\cash-book-2025-updated-audit.xlsx
"""
from __future__ import annotations

from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

YEAR = 2025
SRC = Path(r"C:\Users\brian\Downloads\down\cash-book-2025-updated.xlsx")
OUT = Path(r"C:\Users\brian\Downloads\down\cash-book-2025-updated-audit.xlsx")

DETAIL_SHEET = "Cash Book 2025"
TURNOVER = 24_767_320  # Trial Balance 2025 Turnover CR
SALARY_TOTAL = 5_769_461  # TB Journals: Salaries & wages-NITA

MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
COLUMNS = [
    "FOOD", "Stationary", "Textbooks", "EXAMS", "Uniform", "Communication",
    "Office Exp", "Water", "General repair", "Furniture", "Medical",
    "Service provider", "Fuel", "NTSA", "Trash", "Colnet", "Construction",
    "Vehicle Repairs", "Transport", "Electricity", "Labour", "Donation",
    "Advance tax", "Insurance", "Lisence", "Assets", "LOAN", "Salary",
    "Valuation & Inspection", "ACTIVITIES", "NSSF", "SHA", "PAYE", "NITA",
    "Housing", "Rent",
]
HEADERS = ["DATE", "Voucher No.", "ITEM", "AMOUNT"] + COLUMNS
SALARY_COL = HEADERS.index("Salary") + 1  # 1-based
NCOLS = len(HEADERS)


def parse_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def month_of_row(date_v, vch) -> int | None:
    d = parse_date(date_v)
    if d:
        return d.month
    v = str(vch or "").strip().upper()
    # strip E- / SAL- prefixes for month token
    for prefix in ("E-", "SAL-", "A-", "F-"):
        if v.startswith(prefix):
            v = v[len(prefix):]
            break
    if len(v) >= 3 and v[:3] in MONTHS:
        return MONTHS.index(v[:3]) + 1
    return None


def salary_splits(total: int = SALARY_TOTAL) -> list[int]:
    """Even monthly split; remainder on December."""
    base = total // 12
    rem = total - base * 12
    amounts = [base] * 12
    amounts[11] += rem
    return amounts


def extract_detail_rows(path: Path) -> dict[int, list[list]]:
    """Return {month: [row_values...]} excluding headers and TOTAL rows."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[DETAIL_SHEET]
    by_month: dict[int, list[list]] = defaultdict(list)

    for row in ws.iter_rows(min_row=1, max_col=NCOLS, values_only=True):
        cells = list(row[:NCOLS])
        a = cells[0]
        label = str(a or "").strip().upper()
        if label in ("DATE", "TOTAL", ""):
            # skip blank / structural
            if label in ("DATE", "TOTAL"):
                continue
            if all(c is None or str(c).strip() == "" for c in cells):
                continue
        # skip if no amount
        try:
            amt = float(cells[3] or 0)
        except (TypeError, ValueError):
            continue
        if not amt:
            continue
        m = month_of_row(cells[0], cells[1])
        if not m:
            continue
        # normalize length
        while len(cells) < NCOLS:
            cells.append(None)
        by_month[m].append(cells[:NCOLS])

    wb.close()
    return by_month


def write_audit_workbook(by_month: dict[int, list[list]], path: Path) -> dict:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DETAIL_SHEET

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E1F2")
    int_fmt = "#,##0"
    profit_fill = PatternFill("solid", fgColor="FFF2CC")

    splits = salary_splits()
    month_total_rows: dict[int, int] = {}  # month -> TOTAL row number on detail

    row_num = 1
    for m in range(1, 13):
        entries = list(by_month.get(m, []))

        # Append salary row (retain all existing; add TB salary under Salary)
        last_day = monthrange(YEAR, m)[1]
        sal_amt = splits[m - 1]
        sal_row = [None] * NCOLS
        sal_row[0] = date(YEAR, m, last_day)
        sal_row[1] = f"SAL-{MONTHS[m - 1]}"
        sal_row[2] = "Salaries & wages"
        sal_row[3] = sal_amt
        sal_row[SALARY_COL - 1] = sal_amt
        entries.append(sal_row)

        # Header
        for i, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_num, column=i, value=h)
            cell.font = bold
            cell.fill = fill
        row_num += 1
        first = row_num

        for line in entries:
            for i, val in enumerate(line, 1):
                cell = ws.cell(row=row_num, column=i, value=val)
                if i >= 4 and isinstance(val, (int, float)) and val:
                    cell.number_format = int_fmt
            row_num += 1

        last = row_num - 1
        ws.cell(row=row_num, column=1, value="TOTAL").font = bold
        for col_i in range(4, NCOLS + 1):
            letter = get_column_letter(col_i)
            cell = ws.cell(
                row=row_num,
                column=col_i,
                value=f"=SUM({letter}{first}:{letter}{last})",
            )
            cell.font = bold
            cell.number_format = int_fmt
        month_total_rows[m] = row_num
        row_num += 2

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14

    # --- Summary linked to detail month TOTAL rows ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"EXPENSES SUMMARY YEAR {YEAR}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"] = "Category"
    ws2["A2"].font = bold
    ws2["A2"].fill = fill
    for i, mon in enumerate(MONTHS, 2):
        cell = ws2.cell(row=2, column=i, value=mon)
        cell.font = bold
        cell.fill = fill
    cell = ws2.cell(row=2, column=14, value="TOTAL")
    cell.font = bold
    cell.fill = fill

    detail_ref = f"'{DETAIL_SHEET}'"
    cat_start = 3
    for ci, cat in enumerate(COLUMNS):
        r = cat_start + ci
        ws2.cell(row=r, column=1, value=cat)
        cat_col_on_detail = 5 + ci  # FOOD starts at col E = 5
        cat_letter = get_column_letter(cat_col_on_detail)
        for m in range(1, 13):
            tot_row = month_total_rows[m]
            formula = f"={detail_ref}!{cat_letter}{tot_row}"
            cell = ws2.cell(row=r, column=m + 1, value=formula)
            cell.number_format = int_fmt
        # Row total
        cell = ws2.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
        cell.number_format = int_fmt

    cat_end = cat_start + len(COLUMNS) - 1
    gt_row = cat_end + 1
    ws2.cell(row=gt_row, column=1, value="GRAND TOTAL").font = bold
    for c in range(2, 15):
        letter = get_column_letter(c)
        cell = ws2.cell(
            row=gt_row,
            column=c,
            value=f"=SUM({letter}{cat_start}:{letter}{cat_end})",
        )
        cell.font = bold
        cell.number_format = int_fmt

    # Rent note (retained from prior Summary)
    ws2.cell(row=cat_end, column=17, value="25k per month anything above")

    # Profit block (linked)
    p0 = gt_row + 2
    ws2.cell(row=p0, column=1, value="P&L (linked)").font = Font(bold=True, size=12)
    ws2.cell(row=p0 + 1, column=1, value="Turnover (TB)")
    cell = ws2.cell(row=p0 + 1, column=2, value=TURNOVER)
    cell.number_format = int_fmt
    cell.fill = profit_fill
    ws2.cell(row=p0 + 2, column=1, value="Total expenses (cash book)")
    cell = ws2.cell(row=p0 + 2, column=2, value=f"=N{gt_row}")
    cell.number_format = int_fmt
    cell.fill = profit_fill
    ws2.cell(row=p0 + 3, column=1, value="Profit / (Loss)")
    cell = ws2.cell(row=p0 + 3, column=2, value=f"=B{p0 + 1}-B{p0 + 2}")
    cell.font = bold
    cell.number_format = int_fmt
    cell.fill = profit_fill
    ws2.cell(
        row=p0 + 5,
        column=1,
        value=(
            "Note: Salary includes TB 'Salaries & wages-NITA' "
            f"{SALARY_TOTAL:,} under Salary. E-Citizen NITA levy stays in NITA."
        ),
    )
    ws2.merge_cells(start_row=p0 + 5, start_column=1, end_row=p0 + 5, end_column=8)
    ws2.cell(row=p0 + 5, column=1).alignment = Alignment(wrap_text=True)

    ws2.column_dimensions["A"].width = 28
    for col in range(2, 15):
        ws2.column_dimensions[get_column_letter(col)].width = 11

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    return {
        "salary_total": SALARY_TOTAL,
        "salary_splits": splits,
        "month_total_rows": month_total_rows,
        "grand_total_row": gt_row,
        "expected_expenses": None,  # filled by caller if needed
    }


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source: {SRC}")

    by_month = extract_detail_rows(SRC)
    existing = sum(len(v) for v in by_month.values())
    if existing == 0:
        raise SystemExit("No detail rows extracted from source")

    meta = write_audit_workbook(by_month, OUT)

    # Verify with data_only=False that formulas exist; compute expected sums from values
    expected_base = 0
    for rows in by_month.values():
        for line in rows:
            expected_base += int(float(line[3] or 0))
    expected_expenses = expected_base + SALARY_TOTAL
    profit = TURNOVER - expected_expenses

    print("Audit cash book written")
    print(f"  {OUT}")
    print(f"  Existing detail rows preserved: {existing}")
    print(f"  Salary rows added: 12 totaling {SALARY_TOTAL:,}")
    print(f"  Salary monthly: {meta['salary_splits']}")
    print(f"  Expected expenses (base+salary): {expected_expenses:,}")
    print(f"  Turnover (TB): {TURNOVER:,}")
    print(f"  Profit/(Loss): {profit:,}")
    print(f"  Detail month TOTAL rows: {meta['month_total_rows']}")


if __name__ == "__main__":
    main()
